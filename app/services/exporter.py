from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

UNMATCHED_FILL = PatternFill(fill_type="solid", start_color="FCA5A5", end_color="FCA5A5")


def export_system_rows(
    headers: list[str],
    rows: list[object],
    output_path: str,
    *,
    highlight_unmatched: bool = False,
    sheet_name: str = "HeThong_Xuat",
    attached_statement_path: str | None = None,
    attached_statement_sheet_name: str = "SaoKeGoc",
) -> int:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    exported_rows = 0
    for row in rows:
        sheet.append(row.display_values)
        exported_rows += 1
        if highlight_unmatched and row.status == "unmatched":
            excel_row = sheet.max_row
            for cell in sheet[excel_row]:
                cell.fill = UNMATCHED_FILL
    if exported_rows == 0:
        workbook.close()
        raise ValueError("no-rows")
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)
    if attached_statement_path:
        _append_statement_sheet(workbook, attached_statement_path, attached_statement_sheet_name)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return exported_rows


def export_unmatched_system_rows(result, output_path: str) -> int:
    unmatched_rows = [row for row in result.system_rows if row.status == "unmatched"]
    return export_system_rows(
        result.system_headers,
        unmatched_rows,
        output_path,
        highlight_unmatched=False,
        sheet_name="HeThong_KhongKhop",
    )


def _append_statement_sheet(
    target_workbook: Workbook,
    statement_path: str,
    target_sheet_name: str,
) -> None:
    source_workbook = load_workbook(statement_path)
    try:
        source_sheet = source_workbook.worksheets[0]
        target_sheet = target_workbook.create_sheet(title=target_sheet_name[:31] or "SaoKeGoc")
        for row in source_sheet.iter_rows():
            for source_cell in row:
                target_cell = target_sheet.cell(row=source_cell.row, column=source_cell.column)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.border = copy(source_cell.border)
                    target_cell.alignment = copy(source_cell.alignment)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy(source_cell.comment)
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        for column_key, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[column_key].width = dimension.width
            target_sheet.column_dimensions[column_key].hidden = dimension.hidden
        for row_key, dimension in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_key].height = dimension.height
            target_sheet.row_dimensions[row_key].hidden = dimension.hidden
        target_sheet.freeze_panes = source_sheet.freeze_panes
        if source_sheet.auto_filter:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref
        if source_sheet.sheet_view:
            target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    finally:
        source_workbook.close()
