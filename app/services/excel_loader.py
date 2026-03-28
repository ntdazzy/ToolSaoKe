from __future__ import annotations

import re
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from app.models import BankMetadata, BankTransaction, SystemTransaction
from app.services.utils import (
    amount_to_display,
    compact_spaces,
    contains_tax_keywords,
    extract_reference_prefixes,
    extract_reference_tokens,
    normalize_text,
    parse_date,
    parse_datetime,
    parse_vnd_int,
    to_text,
    tokenize,
)


SYSTEM_HEADERS_LIMIT = 10


def load_system_transactions(path: str) -> tuple[list[str], list[SystemTransaction]]:
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    headers = [
        compact_spaces(to_text(sheet.cell_value(0, col_idx)))
        for col_idx in range(SYSTEM_HEADERS_LIMIT)
    ]
    rows: list[SystemTransaction] = []
    for row_idx in range(1, sheet.nrows):
        raw_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(SYSTEM_HEADERS_LIMIT)]
        summary = compact_spaces(to_text(raw_values[2]))
        if not any(to_text(value) for value in raw_values[:SYSTEM_HEADERS_LIMIT]):
            continue
        if summary == "上期结转":
            continue
        amount_debit = abs(parse_vnd_int(raw_values[4]))
        amount_credit = abs(parse_vnd_int(raw_values[5]))
        direction = "income" if amount_debit > 0 and amount_credit == 0 else "expense"
        amount = amount_debit if direction == "income" else amount_credit
        voucher_date = parse_date(raw_values[0])
        voucher_number = compact_spaces(to_text(raw_values[1]))
        counterpart_account = compact_spaces(to_text(raw_values[3]))
        balance = parse_vnd_int(raw_values[7]) if to_text(raw_values[7]) else None
        data_source = compact_spaces(to_text(raw_values[9]))
        display_values = [
            compact_spaces(to_text(raw_values[0])),
            voucher_number,
            summary,
            counterpart_account,
            amount_to_display(raw_values[4]),
            amount_to_display(raw_values[5]),
            compact_spaces(to_text(raw_values[6])),
            amount_to_display(raw_values[7]),
            compact_spaces(to_text(raw_values[8])),
            data_source,
        ]
        text_source = " ".join(
            value for value in (voucher_number, summary, counterpart_account, data_source) if value
        )
        reference_tokens = extract_reference_tokens(text_source)
        rows.append(
            SystemTransaction(
                row_id=f"sys-{row_idx + 1}",
                excel_row=row_idx + 1,
                display_values=display_values,
                voucher_date=voucher_date,
                voucher_number=voucher_number,
                summary=summary,
                counterpart_account=counterpart_account,
                amount_debit=amount_debit,
                amount_credit=amount_credit,
                direction=direction,
                amount=amount,
                balance=balance,
                data_source=data_source,
                normalized_text=normalize_text(text_source),
                text_tokens=tokenize(text_source),
                reference_tokens=reference_tokens,
                reference_prefixes=extract_reference_prefixes(reference_tokens),
                has_tax=contains_tax_keywords(summary, counterpart_account, data_source),
            )
        )
    return headers, rows


def load_bank_transactions(path: str) -> tuple[list[str], list[BankTransaction], BankMetadata]:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    all_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    metadata = _parse_bank_metadata(all_rows)
    header_idx = _locate_bank_header_row(all_rows)
    headers = [compact_spaces(to_text(cell)) for cell in all_rows[header_idx]]
    last_header_index = max(index for index, value in enumerate(headers) if value)
    headers = headers[: last_header_index + 1]
    transactions: list[BankTransaction] = []
    for row_number, raw_row in enumerate(all_rows[header_idx + 1 :], start=header_idx + 2):
        row = list(raw_row[: len(headers)])
        row.extend([None] * (len(headers) - len(row)))
        if not any(to_text(value) for value in row):
            continue
        request_dt = parse_datetime(row[0])
        transaction_date = parse_date(row[1])
        reference_number = compact_spaces(to_text(row[2]))
        remitter_bank = compact_spaces(to_text(row[3]))
        remitter_account_number = compact_spaces(to_text(row[4]))
        remitter_account_name = compact_spaces(to_text(row[5]))
        description = compact_spaces(to_text(row[6]))
        debit = parse_vnd_int(row[7])
        credit = abs(parse_vnd_int(row[8]))
        fee = parse_vnd_int(row[9])
        vat = parse_vnd_int(row[10])
        running_balance = parse_vnd_int(row[11]) if len(row) > 11 and to_text(row[11]) else None
        expense_total = abs(debit) + abs(fee) + abs(vat)
        direction = "income" if credit > 0 else "expense"
        amount = credit if direction == "income" else expense_total
        display_values = [
            compact_spaces(to_text(row[0])),
            compact_spaces(to_text(row[1])),
            reference_number,
            remitter_bank,
            remitter_account_number,
            remitter_account_name,
            description,
            amount_to_display(row[7]),
            amount_to_display(row[8], blank_for_zero=True),
            amount_to_display(row[9], blank_for_zero=True),
            amount_to_display(row[10], blank_for_zero=True),
            amount_to_display(row[11]) if len(row) > 11 else "",
        ]
        text_source = " ".join(
            value
            for value in (
                reference_number,
                remitter_bank,
                remitter_account_number,
                remitter_account_name,
                description,
            )
            if value
        )
        reference_tokens = extract_reference_tokens(text_source)
        transactions.append(
            BankTransaction(
                row_id=f"bank-{row_number}",
                excel_row=row_number,
                display_values=display_values,
                requesting_datetime=request_dt,
                transaction_date=transaction_date,
                reference_number=reference_number,
                remitter_bank=remitter_bank,
                remitter_account_number=remitter_account_number,
                remitter_account_name=remitter_account_name,
                description=description,
                debit=debit,
                credit=credit,
                fee=fee,
                vat=vat,
                amount=amount,
                direction=direction,
                running_balance=running_balance,
                normalized_text=normalize_text(text_source),
                text_tokens=tokenize(text_source),
                reference_tokens=reference_tokens,
                reference_prefixes=extract_reference_prefixes(reference_tokens),
                has_tax=vat != 0 or contains_tax_keywords(description, reference_number),
            )
        )
    workbook.close()
    return headers, transactions, metadata


def _locate_bank_header_row(rows: list[list[object]]) -> int:
    for index, row in enumerate(rows):
        cells = [compact_spaces(to_text(cell)) for cell in row]
        if (
            "Ngày giao dịch/Transaction date" in cells
            and "Số bút toán/Reference number" in cells
        ):
            return index
    raise ValueError("Khong tim thay dong tieu de cua file sao ke.")


def _parse_bank_metadata(rows: list[list[object]]) -> BankMetadata:
    metadata = BankMetadata()
    if rows:
        metadata.bank_name_vi = compact_spaces(to_text(rows[0][0]))
    if len(rows) > 1:
        metadata.bank_name_en = compact_spaces(to_text(rows[1][0]))
    if len(rows) > 2:
        tax_line = compact_spaces(to_text(rows[2][0]))
        match = re.search(r"MST/Tax code:\s*(.+)$", tax_line)
        if match:
            metadata.tax_code = compact_spaces(match.group(1))
    if len(rows) > 4:
        period_line = " ".join(compact_spaces(to_text(cell)) for cell in rows[4] if to_text(cell))
        period_match = re.search(
            r"From\s+(\d{4}-\d{2}-\d{2}).*To\s+(\d{4}-\d{2}-\d{2})",
            period_line,
            re.IGNORECASE,
        )
        if period_match:
            metadata.from_date = parse_date(period_match.group(1))
            metadata.to_date = parse_date(period_match.group(2))
    label_map = {
        "Số tài khoản/Account number": ("account_number", False),
        "Tên tài khoản/Account name": ("account_name", False),
        "Loại tiền/Currency": ("currency", False),
        "Loại tài khoản/Account type": ("account_type", False),
        "Số dư hiện tại/Actual balance": ("actual_balance", True),
        "Số dư đầu ngày/Opening balance": ("opening_balance", True),
        "Số dư cuối ngày/ Closing balance": ("closing_balance", True),
        "Tổng ghi nợ/ Total debits": ("total_debits", True),
        "Tổng ghi có/ Total credits": ("total_credits", True),
        "Tổng phí/ Total fees": ("total_fees", True),
        "Tổng VAT/ Total VAT": ("total_vat", True),
        "Tổng lệnh ghi nợ/ Total debit transaction": ("total_debit_transactions", True),
        "Tổng lệnh ghi có/ Total credit transaction": ("total_credit_transactions", True),
    }
    for row in rows[:18]:
        cleaned = [compact_spaces(to_text(cell)) for cell in row]
        for index, value in enumerate(cleaned):
            if not value or value not in label_map:
                continue
            field_name, is_number = label_map[value]
            next_value = _next_non_empty(cleaned, index + 1)
            setattr(metadata, field_name, parse_vnd_int(next_value) if is_number else next_value)
    return metadata


def _next_non_empty(values: list[str], start_index: int) -> str:
    for value in values[start_index:]:
        if value:
            return value
    return ""


def is_supported_excel_file(path: str) -> bool:
    return Path(path).suffix.lower() in {".xls", ".xlsx"}
